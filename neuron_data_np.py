"""
Convert the neuropeptide connectivity spreadsheet into an edge list between neuron classes
"""


from openpyxl import load_workbook
from collections import defaultdict, namedtuple
from os.path import join
import json
import csv

with open('paper_data/urls2.json') as f:
    urls = json.load(f)
with open('paper_data/names.json') as f:
    names = json.load(f)

src_root = '/home/cbarnes/work/code/connectome/construct2/extrasyn/src_data/'

LATEST_DATA = join(src_root, 'neuropeptide_spreadsheet.xlsx')

Edge = namedtuple('Edge', ['src', 'tgt', 'transmitter', 'receptor'])

pep_sht_name = 'Peptide Expr'
rec_sht_name = 'Receptor Expr'
mapping_sht_name = 'Peptides'


def val(cell):
    try:
        return cell.value.strip()
    except AttributeError as e:
        if 'NoneType' in str(e):
            return ''
        else:
            raise e


def peptide_expr_from_sheet(sheet):
    rows = []
    for row in sheet.iter_rows('A3:D{}'.format(sheet.get_highest_row())):
        if not row:
            continue
        if val(row[0]):
            if '?' in val(row[1]):
                continue

            peptide = val(row[0])
            if 'nlp-37' in peptide:
                peptide = 'pdf-2'

            rows.append([
                val(row[1]).replace('?', ''),
                'Neuropeptide',
                peptide.upper(),
                names[val(row[3])],
                urls[val(row[3])]
            ])

    return sorted(rows)


def receptor_expr_from_sheet(sheet):
    rows = []
    for row in sheet.iter_rows('A3:D{}'.format(sheet.get_highest_row())):
        if not row:
            continue
        if val(row[0]):
            if '?' in val(row[1]):
                continue

            receptor = val(row[0])

            rows.append([
                val(row[1]).replace('?', ''),
                'Receptor',
                receptor.upper(),
                names[val(row[3])],
                urls[val(row[3])]
            ])

    return sorted(rows)


def ligand_mapping_from_sheet(sheet):
    rows = []
    for row in sheet.iter_rows('A3:J{}'.format(sheet.get_highest_row())):
        if not row:
            continue
        if val(row[0]) and val(row[1]):
            if '?' in val(row[1]):
                continue

            pep = val(row[0])

            if 'nlp-37' in pep:
                pep = 'pdf-2'

            for rec_cell in row[1:]:
                if not rec_cell or not val(rec_cell):
                    break

                rows.append([
                    pep.upper(),
                    'Ligand',
                    val(rec_cell).upper(),
                    'UNKNOWN',
                    'UNKNOWN'
                ])

    return sorted(rows)


def main():
    wb = load_workbook(LATEST_DATA, read_only=True)

    pep_sht = wb.get_sheet_by_name(pep_sht_name)
    pep_sht.calculate_dimension(force=True)
    peptide_expr = peptide_expr_from_sheet(pep_sht)

    rec_sht = wb.get_sheet_by_name(rec_sht_name)
    rec_sht.calculate_dimension(force=True)
    rec_expr = receptor_expr_from_sheet(rec_sht)

    mapping_sht = wb.get_sheet_by_name(mapping_sht_name)
    mapping_sht.calculate_dimension(force=True)
    ligand_mapping = ligand_mapping_from_sheet(mapping_sht)

    with open('np_exprs.csv', 'w') as f:
        writer = csv.writer(f)
        writer.writerow(['Entity1', 'Relationship', 'Entity2', 'Evidence', 'Evidence URL'])
        writer.writerows(sorted(peptide_expr))
        writer.writerows(sorted(rec_expr))
        writer.writerows(sorted(ligand_mapping))

if __name__ == '__main__':
    main()
    print('done')