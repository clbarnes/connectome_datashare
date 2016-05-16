from openpyxl import load_workbook
from collections import defaultdict, namedtuple
from itertools import product, chain
from os.path import join
import json
import csv
# from extrasyn.paths import src_root, tgt_root

src_root = '/home/cbarnes/work/code/connectome/construct2/extrasyn/src_data/'

LATEST_DATA = join(src_root, 'monoamine_spreadsheet.xlsx')


monoamines = ['serotonin', 'dopamine', 'octopamine', 'tyramine']

ma_sht_name = 'Monoamine Expr'
rec_sht_name = 'Receptor Expr'
ligand_sht_names = ['5-HT', 'DA', 'OA', 'TA']

with open('paper_data/urls2.json') as f:
    urls = json.load(f)
with open('paper_data/names.json') as f:
    names = json.load(f)


def val(cell):
    try:
        return cell.value.strip()
    except AttributeError as e:
        if 'NoneType' in str(e):
            return ''
        else:
            raise e


def ma_cell_sheet_to_rows(sheet, include_weak=False):
    rows = []
    for row in sheet.iter_rows('A3:E{}'.format(sheet.max_row)):
        if not row:
            continue
        if row[0].value:
            if '*' in row[2].value:
                continue
            if '?' in row[2].value and not include_weak:
                continue

            rows.append([
                val(row[2]).replace('?', ''),
                'Neurotransmitter',
                val(row[0]).title(),
                names[val(row[4])],
                urls[val(row[4])]
            ])

    return rows


def rec_cell_sheet_to_exprs(sheet):
    rec_expr_rows = []
    for row in sheet.iter_rows('A3:E{}'.format(sheet.max_row)):
        if not row:
            continue

        if row[0].value:
            transmitter, receptor, cell, _, citation = [val(item) for item in row[:5]]

            rec_expr_rows.append([
                cell,
                'Receptor',
                receptor.upper(),
                names.get(citation, 'Schafer et al., (n.d.)'),
                urls.get(citation, 'THIS_STUDY')
            ])

    return rec_expr_rows


def all_ligand_rows(wb):
    lol = []
    for ligand_sht_name in ligand_sht_names:
        sheet = get_sheet(wb, ligand_sht_name)
        lol.append(ligand_sheet_to_rows(sheet))

    return list(chain(*lol))


def ligand_sheet_to_rows(sheet):
    ligand_rows = []
    for row in sheet.iter_rows('A3:H{}'.format(sheet.max_row)):
        if not row:
            continue

        if val(row[1]) != 'receptor':
            continue

        transmitter, _is_rec, receptor, _wbid, _cls, _note, _hom, reference = [val(item) for item in row[:8]]

        ligand_rows.append([transmitter.title(), 'Ligand', receptor.upper(), names[reference], urls[reference]])

    return ligand_rows


def get_sheet(workbook, name):
    sheet = workbook.get_sheet_by_name(name)
    print(name)
    # sheet.calculate_dimension(force=True)
    return sheet


def main(include_weak):
    wb = load_workbook(LATEST_DATA)
    ma_sht = get_sheet(wb, ma_sht_name)
    rec_sht = get_sheet(wb, rec_sht_name)

    ma_expr = ma_cell_sheet_to_rows(ma_sht, include_weak=include_weak)
    ma_rec_expr = rec_cell_sheet_to_exprs(rec_sht)

    ligand_info = all_ligand_rows(wb)

    with open('ma_exprs.csv', 'w') as f:
        writer = csv.writer(f)
        writer.writerow(['Entity1', 'Relationship', 'Entity2', 'Evidence', 'Evidence URL'])
        writer.writerows(sorted(ma_expr))
        writer.writerows(sorted(ma_rec_expr))
        writer.writerows(sorted(ligand_info))


if __name__ == '__main__':
    main(include_weak=False)
    print('done')